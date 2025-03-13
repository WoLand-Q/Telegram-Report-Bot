import logging
import sys
import os
import glob
import datetime
import requests
import openpyxl
import json
from datetime import time
import pytz
from typing import Dict, List, Any, Tuple
from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ConversationHandler,
    filters,
    ContextTypes,
)

# ----------------- ЛОГИРОВАНИЕ -----------------
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)

# ----------------- НАСТРОЙКИ -----------------
BOT_TOKEN = ""  # Ваш токен
ADMIN_CHAT_ID =  43545 # ID чата для отчётов

IIKO_HOST = ""
LOGIN = ""
PASSWORD_SHA1 = ""  # SHA1

PLAN_FACT_FOLDER = "data_excels"

# Используемые категории
CATEGORIES = ["доставка", "зал", "агрегаторы"]

REPORT_TYPE = "SALES"
GROUP_BY_ROW_FIELDS = ["Department", "OrderType"]
AGGREGATE_FIELDS = [
    "GuestNum",
    "UniqOrderId.OrdersCount",
    "DishDiscountSumInt",
    "DishDiscountSumInt.average"
]
BUILD_SUMMARY = True

FILTERS = {
    "OpenDate.Typed": {
        "filterType": "DateRange",
        "periodType": "CUSTOM",
        "from": None,
        "to": None,
        "includeLow": True,
        "includeHigh": False
    },
    "DeletedWithWriteoff": {"filterType": "IncludeValues", "values": ["NOT_DELETED"]},
    "OrderDeleted": {"filterType": "IncludeValues", "values": ["NOT_DELETED"]}
}

# ----- Новые настройки для автоотчёта -----
# Определяем сети: ключ – имя сети, значение – список точек (имён файлов без расширения)
NETWORK_GROUPS = {
    "Киев": [""],
    "Днепр": ["", "", ""],
    "Харьков": ["", "", ""]
}
# Файл с Telegram ID для автоотчётов (например, [123456789, 987654321])
AUTO_REPORT_USERS_FILE = "auto_report_users.json"


# ----------- Функция экранирования Markdown -----------
def escape_markdown(text: str) -> str:
    """
    Экранирует спецсимволы Markdown (классический Markdown) – *, _, `, [.
    """
    text = text.replace("\\", "\\\\")
    escape_chars = r"*_[`["
    for char in escape_chars:
        text = text.replace(char, f"\\{char}")
    return text


# ----------------- ФУНКЦИИ ДЛЯ IIKO -----------------
def map_order_type_to_category(order_type_str: str) -> str:
    if not order_type_str:
        return "зал"
    lower_ot = order_type_str.lower()
    if any(x in lower_ot for x in ["bolt", "glovo", "delivery hub", "пюрешка & котлетка"]):
        return "агрегаторы"
    return "доставка"


def safe_float(cell_value) -> float:
    if cell_value is None:
        return 0.0
    try:
        val_str = str(cell_value).replace('\xa0', '').strip()
        return float(val_str) if val_str else 0.0
    except ValueError:
        return 0.0


def iiko_login(session: requests.Session) -> str:
    auth_url = f"{IIKO_HOST}/resto/api/auth"
    payload = {"login": LOGIN, "pass": PASSWORD_SHA1}
    try:
        resp = session.post(auth_url, data=payload, timeout=10)
        resp.raise_for_status()
        token = resp.text.strip()
        logging.info("Успешная авторизация в iiko. Токен: %s", token)
        return token
    except requests.exceptions.RequestException as exc:
        logging.error("Ошибка при авторизации iiko: %s", exc)
        sys.exit(1)


def iiko_logout(session: requests.Session, token: str):
    logout_url = f"{IIKO_HOST}/resto/api/logout"
    payload = {"key": token}
    try:
        resp = session.post(logout_url, data=payload, timeout=10)
        if resp.status_code == 200:
            logging.info("Успешный logout в iiko.")
        else:
            logging.warning("Logout вернул статус %s", resp.status_code)
    except requests.exceptions.RequestException as exc:
        logging.warning("Ошибка при logout iiko: %s", exc)


def build_olap_request_body(filters: dict) -> dict:
    return {
        "reportType": REPORT_TYPE,
        "buildSummary": BUILD_SUMMARY,
        "groupByRowFields": GROUP_BY_ROW_FIELDS,
        "aggregateFields": AGGREGATE_FIELDS,
        "filters": filters
    }


def fetch_olap_report(session: requests.Session, token: str, body: dict) -> dict:
    olap_url = f"{IIKO_HOST}/resto/api/v2/reports/olap"
    headers = {"Content-Type": "application/json; charset=utf-8"}
    try:
        resp = session.post(
            olap_url,
            params={"key": token},
            headers=headers,
            json=body,
            timeout=30
        )
        resp.raise_for_status()
        data = resp.json()
        logging.info("OLAP ответ:\n%s", json.dumps(data, ensure_ascii=False, indent=4))
        return data
    except requests.exceptions.RequestException as exc:
        logging.error("Ошибка при получении OLAP: %s", exc)
        sys.exit(1)


def get_report_for_department(session: requests.Session, token: str,
                              department_name: str, date_from: str, date_to: str) -> list:
    filters_updated = FILTERS.copy()
    filters_updated["OpenDate.Typed"]["from"] = f"{date_from}T00:00:00.000"
    filters_updated["OpenDate.Typed"]["to"] = f"{date_to}T00:00:00.000"
    filters_updated["Department"] = {"filterType": "IncludeValues", "values": [department_name]}
    body = build_olap_request_body(filters_updated)
    res_json = fetch_olap_report(session, token, body)
    data_rows = res_json.get("data", [])
    logging.info("Для заведения '%s' получено %d строк из OLAP.", department_name, len(data_rows))
    return data_rows


# ---------------- Функции для работы с план/факт из Excel ----------------
def parse_plan_fact_excel(file_path: str) -> Dict[Tuple[str, str], Dict[str, float]]:
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    plan_fact_data = {}
    for row in range(2, sheet.max_row + 1):
        raw_date = sheet.cell(row=row, column=2).value
        if not raw_date:
            continue
        try:
            dt = datetime.datetime.strptime(str(raw_date).strip(), "%d.%m.%Y")
            date_key = dt.date().isoformat()
        except Exception as e:
            logging.error("Ошибка преобразования даты '%s': %s", raw_date, e)
            continue

        plan_total_sales = safe_float(sheet.cell(row=row, column=3).value)
        plan_sales_hall = safe_float(sheet.cell(row=row, column=4).value)
        plan_sales_deliv = safe_float(sheet.cell(row=row, column=5).value)
        plan_sales_agg = safe_float(sheet.cell(row=row, column=6).value)
        plan_avg_check_hall = safe_float(sheet.cell(row=row, column=7).value)
        plan_avg_guest_hall = safe_float(sheet.cell(row=row, column=8).value)  # не используется
        plan_guests_hall = safe_float(sheet.cell(row=row, column=9).value)  # "Гости Зал"
        plan_orders_hall = safe_float(sheet.cell(row=row, column=10).value)
        plan_avg_check_deliv = safe_float(sheet.cell(row=row, column=11).value)
        plan_avg_check_agg = safe_float(sheet.cell(row=row, column=12).value)
        plan_orders_deliv = safe_float(sheet.cell(row=row, column=13).value)
        plan_orders_agg = safe_float(sheet.cell(row=row, column=14).value)

        plan_fact_data[(date_key, "итого")] = {
            "plan_total_sales": plan_total_sales
        }
        plan_fact_data[(date_key, "зал")] = {
            "plan_sales": plan_sales_hall,
            "plan_orders": plan_orders_hall,
            "plan_avg_check": plan_avg_check_hall,
            "plan_guests": plan_guests_hall
        }
        plan_fact_data[(date_key, "доставка")] = {
            "plan_sales": plan_sales_deliv,
            "plan_orders": plan_orders_deliv,
            "plan_avg_check": plan_avg_check_deliv,
        }
        plan_fact_data[(date_key, "агрегаторы")] = {
            "plan_sales": plan_sales_agg,
            "plan_orders": plan_orders_agg,
            "plan_avg_check": plan_avg_check_agg,
        }
    return plan_fact_data


def combine_plan_fact_with_iiko(
        plan_fact_data: Dict[Tuple[str, str], Dict[str, float]],
        iiko_data: List[Dict[str, Any]],
        target_date: str,
        category: str
) -> Dict[str, float]:
    key = (target_date, category.lower())
    pf_values = plan_fact_data.get(key, {})

    fact_sum = 0.0
    fact_orders = 0.0
    fact_guests = 0.0

    for row in iiko_data:
        order_type_raw = (row.get("OrderType") or "").strip()
        cat_mapped = map_order_type_to_category(order_type_raw)
        if cat_mapped == category.lower():
            fact_sum += float(row.get("DishDiscountSumInt", 0.0))
            fact_orders += float(row.get("UniqOrderId.OrdersCount", 0.0))
            if category.lower() == "зал":
                fact_guests += float(row.get("GuestNum", 0.0))

    combined = {
        "plan_sales": pf_values.get("plan_sales", 0.0),
        "plan_orders": pf_values.get("plan_orders", 0.0),
        "plan_avg_check": pf_values.get("plan_avg_check", 0.0),
        "fact_sales": fact_sum,
        "fact_orders": fact_orders,
        "fact_avg_check": fact_sum / fact_orders if fact_orders else 0.0
    }
    if category.lower() == "зал":
        combined["plan_guests"] = pf_values.get("plan_guests", 0.0)
        combined["fact_guests"] = fact_guests
    return combined


def get_detailed_plan_fact(department: str, target_date: str) -> Dict[str, Any]:
    """
    Получает подробный план/факт для заведения (имя файла = название заведения)
    за указанную дату (формат YYYY-MM-DD) с разбивкой по категориям и общей сводкой.
    """
    file_path = os.path.join(PLAN_FACT_FOLDER, f"{department}.xlsx")
    if not os.path.exists(file_path):
        logging.error("Файл для заведения '%s' не найден.", department)
        return {}
    pf_data = parse_plan_fact_excel(file_path)
    date_from = target_date
    date_to = (datetime.datetime.strptime(target_date, "%Y-%m-%d") + datetime.timedelta(days=1)).date().isoformat()

    with requests.Session() as session:
        token = iiko_login(session)
        iiko_data = get_report_for_department(session, token, department, date_from, date_to)
        iiko_logout(session, token)

    details = {}
    overall_fact_sales = 0.0
    overall_fact_orders = 0.0
    overall_plan_orders = 0.0
    overall_fact_guests = 0.0

    for cat in CATEGORIES:
        res = combine_plan_fact_with_iiko(pf_data, iiko_data, target_date, cat)
        details[cat] = res
        overall_fact_sales += res["fact_sales"]
        overall_fact_orders += res["fact_orders"]
        overall_plan_orders += res["plan_orders"]
        if cat.lower() == "зал":
            overall_fact_guests += res.get("fact_guests", 0.0)

    overall_plan = pf_data.get((target_date, "итого"), {}).get("plan_total_sales", 0.0)
    overall_plan_avg = overall_plan / overall_plan_orders if overall_plan_orders else 0.0
    overall_fact_avg = overall_fact_sales / overall_fact_orders if overall_fact_orders else 0.0

    overall = {
        "plan_total_sales": overall_plan,
        "plan_orders": overall_plan_orders,
        "fact_sales": overall_fact_sales,
        "fact_orders": overall_fact_orders,
        "plan_avg_check": overall_plan_avg,
        "fact_avg_check": overall_fact_avg,
        "plan_guests": pf_data.get((target_date, "зал"), {}).get("plan_guests", 0.0),
        "fact_guests": overall_fact_guests
    }

    return {
        "department": department,
        "target_date": target_date,
        "details": details,
        "overall": overall
    }


# ----------------- Функция для отправки длинного сообщения -----------------
async def send_long_message(context: ContextTypes.DEFAULT_TYPE, chat_id: int, text: str, max_length: int = 3500):
    """
    Отправляет сообщение частями, разбивая по строкам, чтобы не разрывать Markdown-сущности.
    """
    lines = text.split("\n")
    chunk = ""
    for line in lines:
        if len(chunk) + len(line) + 1 > max_length:
            await context.bot.send_message(chat_id=chat_id, text=chunk, parse_mode="Markdown")
            chunk = line
        else:
            if chunk:
                chunk += "\n" + line
            else:
                chunk = line
    if chunk:
        await context.bot.send_message(chat_id=chat_id, text=chunk, parse_mode="Markdown")


# ----------------- Интерфейс /get_plan_fact через ConversationHandler -----------------
GET_DATE, CHOOSE_DEPARTMENT = range(2)


async def get_plan_fact_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Введите дату (в формате YYYY-MM-DD) для получения плана/факта:")
    return GET_DATE


async def get_date_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    try:
        dt = datetime.datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        await update.message.reply_text("Неверный формат даты. Пожалуйста, введите дату в формате YYYY-MM-DD:")
        return GET_DATE
    context.user_data["target_date"] = dt.isoformat()

    # Список заведений — имена файлов без расширения
    files = glob.glob(os.path.join(PLAN_FACT_FOLDER, "*.xlsx"))
    if not files:
        await update.message.reply_text("Нет загруженных файлов с данными заведений.")
        return ConversationHandler.END
    departments = [os.path.splitext(os.path.basename(f))[0] for f in files]

    keyboard = [[InlineKeyboardButton(dept, callback_data=dept)] for dept in departments]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Выберите заведение:", reply_markup=reply_markup)
    return CHOOSE_DEPARTMENT


async def choose_department_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    department = query.data
    target_date = context.user_data.get("target_date")
    if not target_date:
        await query.edit_message_text("Ошибка: не задана дата.")
        return ConversationHandler.END

    data = get_detailed_plan_fact(department, target_date)
    if not data:
        await query.edit_message_text("Нет данных для заданных параметров.")
        return ConversationHandler.END

    # Формируем сообщение с детализацией по категориям и общей сводкой
    emoji_map = {"доставка": "🚚", "зал": "🏰", "агрегаторы": "📦"}
    lines = []
    lines.append(f"🏢 Заведение: {escape_markdown(department)}")
    lines.append("---\n")
    details = data["details"]
    for cat in CATEGORIES:
        cat_data = details.get(cat, {})
        emoji = emoji_map.get(cat.lower(), "•")
        cat_title = cat.capitalize()
        lines.append(f"{emoji} {escape_markdown(cat_title)}:")
        lines.append(
            f"• *План Продажи:* {cat_data.get('plan_sales', 0):.0f} грн | *Факт Продажи:* {cat_data.get('fact_sales', 0):.0f} грн")
        lines.append(
            f"• *План Заказов:* {cat_data.get('plan_orders', 0):.0f} | *Факт Заказов:* {cat_data.get('fact_orders', 0):.0f}")
        lines.append(
            f"• *План Ср.Заказ:* {cat_data.get('plan_avg_check', 0):.0f} грн | *Факт Ср.Заказ:* {cat_data.get('fact_avg_check', 0):.2f} грн")
        if cat.lower() == "зал":
            lines.append(
                f"• *План Гостей Зал:* {cat_data.get('plan_guests', 0):.0f} | *Факт Гостей:* {cat_data.get('fact_guests', 0):.0f}")
        lines.append("")
    overall = data["overall"]
    lines.append("🏷️ Общая (доставка+зал+агрегаторы):")
    lines.append(
        f"• *План Продажи (Итого):* {overall.get('plan_total_sales', 0):.0f} грн | *Факт Продажи:* {overall.get('fact_sales', 0):.0f} грн")
    lines.append(
        f"• *План Заказов (сумм.):* {overall.get('plan_orders', 0):.0f} | *Факт Заказов:* {overall.get('fact_orders', 0):.0f}")
    lines.append(
        f"• *План Ср.Заказ:* {overall.get('plan_avg_check', 0):.0f} грн | *Факт Ср.Заказ:* {overall.get('fact_avg_check', 0):.2f} грн")
    lines.append(
        f"• *План Гостей (зал):* {overall.get('plan_guests', 0):.0f} | *Факт Гостей:* {overall.get('fact_guests', 0):.0f}")
    final_text = "\n".join(lines)
    await query.edit_message_text(final_text, parse_mode="Markdown")
    return ConversationHandler.END


async def cancel_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Операция отменена.")
    return ConversationHandler.END


# ----------------- Новая функция: Агрегированный автоотчёт по сетям -----------------
def get_aggregated_network_plan_fact(target_date: str) -> Dict[str, Any]:
    """
    Для заданной даты (YYYY-MM-DD) получает агрегированные данные по всем точкам, входящим в сети,
    заданные в NETWORK_GROUPS. Возвращает словарь с агрегированными данными по категориям и общую сводку.
    """
    # Инициализируем аккумуляторы для каждой категории
    agg_categories = {cat: {"plan_sales": 0.0, "fact_sales": 0.0,
                            "plan_orders": 0.0, "fact_orders": 0.0} for cat in CATEGORIES}
    # Для зала добавляем гостей
    agg_categories["зал"].update({"plan_guests": 0.0, "fact_guests": 0.0})
    # Общие аккумуляторы
    overall = {"plan_total_sales": 0.0, "plan_orders": 0.0,
               "fact_sales": 0.0, "fact_orders": 0.0,
               "plan_guests": 0.0, "fact_guests": 0.0}

    # Перебираем сети
    for network, departments in NETWORK_GROUPS.items():
        for dept in departments:
            file_path = os.path.join(PLAN_FACT_FOLDER, f"{dept}.xlsx")
            if not os.path.exists(file_path):
                logging.warning("Файл для точки '%s' не найден, пропускаем.", dept)
                continue
            data = get_detailed_plan_fact(dept, target_date)
            if not data:
                continue
            # По категориям
            details = data.get("details", {})
            for cat in CATEGORIES:
                cat_data = details.get(cat, {})
                agg_categories[cat]["plan_sales"] += cat_data.get("plan_sales", 0)
                agg_categories[cat]["fact_sales"] += cat_data.get("fact_sales", 0)
                agg_categories[cat]["plan_orders"] += cat_data.get("plan_orders", 0)
                agg_categories[cat]["fact_orders"] += cat_data.get("fact_orders", 0)
                if cat.lower() == "зал":
                    agg_categories[cat]["plan_guests"] += cat_data.get("plan_guests", 0)
                    agg_categories[cat]["fact_guests"] += cat_data.get("fact_guests", 0)
            # Общая сводка
            overall_data = data.get("overall", {})
            overall["plan_total_sales"] += overall_data.get("plan_total_sales", 0)
            overall["plan_orders"] += overall_data.get("plan_orders", 0)
            overall["fact_sales"] += overall_data.get("fact_sales", 0)
            overall["fact_orders"] += overall_data.get("fact_orders", 0)
            overall["plan_guests"] += overall_data.get("plan_guests", 0)
            overall["fact_guests"] += overall_data.get("fact_guests", 0)

    # Вычисляем средние (средний чек) для каждой категории
    for cat in CATEGORIES:
        cat_dict = agg_categories[cat]
        if cat_dict["plan_orders"]:
            cat_dict["plan_avg_check"] = cat_dict["plan_sales"] / cat_dict["plan_orders"]
        else:
            cat_dict["plan_avg_check"] = 0.0
        if cat_dict["fact_orders"]:
            cat_dict["fact_avg_check"] = cat_dict["fact_sales"] / cat_dict["fact_orders"]
        else:
            cat_dict["fact_avg_check"] = 0.0

    # Общий средний чек (суммируем по всем категориям)
    total_plan_orders = sum(agg_categories[cat]["plan_orders"] for cat in CATEGORIES)
    total_fact_orders = sum(agg_categories[cat]["fact_orders"] for cat in CATEGORIES)
    if total_plan_orders:
        overall["plan_avg_check"] = sum(agg_categories[cat]["plan_sales"] for cat in CATEGORIES) / total_plan_orders
    else:
        overall["plan_avg_check"] = 0.0
    if total_fact_orders:
        overall["fact_avg_check"] = sum(agg_categories[cat]["fact_sales"] for cat in CATEGORIES) / total_fact_orders
    else:
        overall["fact_avg_check"] = 0.0

    return {
        "networks": list(NETWORK_GROUPS.keys()),
        "categories": agg_categories,
        "overall": overall
    }


async def auto_report_job(context: ContextTypes.DEFAULT_TYPE):
    """
    Ежедневная задача, которая отправляет автоотчёт за предыдущий день (агрегированный по всем сетям)
    на список Telegram-ID, указанных в файле AUTO_REPORT_USERS_FILE.
    """
    # Определяем дату предыдущего дня
    yesterday = datetime.date.today() - datetime.timedelta(days=1)
    target_date = yesterday.isoformat()

    # Получаем агрегированные данные
    agg_data = get_aggregated_network_plan_fact(target_date)
    networks = ", ".join(agg_data["networks"])
    cat_data = agg_data["categories"]
    overall = agg_data["overall"]

    # Формируем текст отчёта
    lines = []
    lines.append(f"*Автоотчёт за {escape_markdown(target_date)}*")
    lines.append(f"\nСеть: {escape_markdown(networks)}")
    lines.append("---\n")

    # Для каждой категории
    emoji_map = {"доставка": "🚚", "зал": "🏰", "агрегаторы": "📦"}
    for cat in CATEGORIES:
        data_cat = cat_data.get(cat, {})
        emoji = emoji_map.get(cat.lower(), "•")
        cat_title = cat.capitalize()
        lines.append(f"{emoji} {escape_markdown(cat_title)}:")
        lines.append(
            f"• *План Продажи:* {data_cat.get('plan_sales', 0):.0f} грн | *Факт Продажи:* {data_cat.get('fact_sales', 0):.0f} грн")
        lines.append(
            f"• *План Заказов:* {data_cat.get('plan_orders', 0):.0f} | *Факт Заказов:* {data_cat.get('fact_orders', 0):.0f}")
        lines.append(
            f"• *План Ср.Заказ:* {data_cat.get('plan_avg_check', 0):.0f} грн | *Факт Ср.Заказ:* {data_cat.get('fact_avg_check', 0):.2f} грн")
        if cat.lower() == "зал":
            lines.append(
                f"• *План Гостей Зал:* {data_cat.get('plan_guests', 0):.0f} | *Факт Гостей:* {data_cat.get('fact_guests', 0):.0f}")
        lines.append("")

    # Общая сводка
    lines.append("🏷️ Общая (доставка+зал+агрегаторы):")
    lines.append(
        f"• *План Продажи (Итого):* {overall.get('plan_total_sales', 0):.0f} грн | *Факт Продажи:* {overall.get('fact_sales', 0):.0f} грн")
    lines.append(
        f"• *План Заказов (сумм.):* {overall.get('plan_orders', 0):.0f} | *Факт Заказов:* {overall.get('fact_orders', 0):.0f}")
    lines.append(
        f"• *План Ср.Заказ:* {overall.get('plan_avg_check', 0):.0f} грн | *Факт Ср.Заказ:* {overall.get('fact_avg_check', 0):.2f} грн")
    lines.append(
        f"• *План Гостей (зал):* {overall.get('plan_guests', 0):.0f} | *Факт Гостей:* {overall.get('fact_guests', 0):.0f}")

    final_text = "\n".join(lines)

    # Читаем список Telegram ID для автоотчётов из JSON-файла
    if os.path.exists(AUTO_REPORT_USERS_FILE):
        try:
            with open(AUTO_REPORT_USERS_FILE, "r", encoding="utf-8") as f:
                user_ids = json.load(f)
        except Exception as e:
            logging.error("Ошибка чтения файла %s: %s", AUTO_REPORT_USERS_FILE, e)
            return
    else:
        logging.warning("Файл %s не найден. Автоотчет не отправлен.", AUTO_REPORT_USERS_FILE)
        return

    # Отправляем отчёт каждому пользователю из списка
    for uid in user_ids:
        try:
            await context.bot.send_message(chat_id=uid, text=final_text, parse_mode="Markdown")
        except Exception as e:
            logging.error("Ошибка отправки автоотчёта пользователю %s: %s", uid, e)

    logging.info("Автоотчёт за %s успешно отправлен.", target_date)


# ----------------- Остальные команды бота -----------------
async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Привет! Я бот для автоматического отчета.")


async def upload_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Пришлите .xlsx-файл — он сохранится в папку data_excels.")


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    document = update.message.document
    file_name = document.file_name
    if file_name.endswith(".xlsx"):
        file_path = os.path.join(PLAN_FACT_FOLDER, file_name)
        file_obj = await document.get_file()
        await file_obj.download_to_drive(file_path)
        await update.message.reply_text(f"Файл '{file_name}' сохранён.")
    else:
        await update.message.reply_text("Это не .xlsx-файл.")


async def test_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Для теста выбираем первое заведение из папки
    files = glob.glob(os.path.join(PLAN_FACT_FOLDER, "*.xlsx"))
    if not files:
        await update.message.reply_text("Нет загруженных файлов для теста.")
        return
    department = os.path.splitext(os.path.basename(files[0]))[0]
    target_date = datetime.date.today().isoformat()
    data = get_detailed_plan_fact(department, target_date)
    if not data:
        await update.message.reply_text("Нет данных для теста.")
        return

    emoji_map = {"доставка": "🚚", "зал": "🏰", "агрегаторы": "📦"}
    lines = []
    lines.append(f"🏢 Заведение: {escape_markdown(department)}")
    lines.append("---\n")
    details = data["details"]
    for cat in CATEGORIES:
        cat_data = details.get(cat, {})
        emoji = emoji_map.get(cat.lower(), "•")
        cat_title = cat.capitalize()
        lines.append(f"{emoji} {escape_markdown(cat_title)}:")
        lines.append(
            f"• *План Продажи:* {cat_data.get('plan_sales', 0):.0f} грн | *Факт Продажи:* {cat_data.get('fact_sales', 0):.0f} грн")
        lines.append(
            f"• *План Заказов:* {cat_data.get('plan_orders', 0):.0f} | *Факт Заказов:* {cat_data.get('fact_orders', 0):.0f}")
        lines.append(
            f"• *План Ср.Заказ:* {cat_data.get('plan_avg_check', 0):.0f} грн | *Факт Ср.Заказ:* {cat_data.get('fact_avg_check', 0):.2f} грн")
        if cat.lower() == "зал":
            lines.append(
                f"• *План Гостей Зал:* {cat_data.get('plan_guests', 0):.0f} | *Факт Гостей:* {cat_data.get('fact_guests', 0):.0f}")
        lines.append("")
    overall = data["overall"]
    lines.append("🏷️ Общая (доставка+зал+агрегаторы):")
    lines.append(
        f"• *План Продажи (Итого):* {overall.get('plan_total_sales', 0):.0f} грн | *Факт Продажи:* {overall.get('fact_sales', 0):.0f} грн")
    lines.append(
        f"• *План Заказов (сумм.):* {overall.get('plan_orders', 0):.0f} | *Факт Заказов:* {overall.get('fact_orders', 0):.0f}")
    lines.append(
        f"• *План Ср.Заказ:* {overall.get('plan_avg_check', 0):.0f} грн | *Факт Ср.Заказ:* {overall.get('fact_avg_check', 0):.2f} грн")
    lines.append(
        f"• *План Гостей (зал):* {overall.get('plan_guests', 0):.0f} | *Факт Гостей:* {overall.get('fact_guests', 0):.0f}")
    final_text = "\n".join(lines)
    await send_long_message(context, update.effective_chat.id, final_text)


def main():
    os.makedirs(PLAN_FACT_FOLDER, exist_ok=True)
    app = ApplicationBuilder().token(BOT_TOKEN).build()

    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("get_plan_fact", get_plan_fact_start)],
        states={
            GET_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_date_handler)],
            CHOOSE_DEPARTMENT: [CallbackQueryHandler(choose_department_handler)]
        },
        fallbacks=[CommandHandler("cancel", cancel_handler)],
    )

    app.add_handler(CommandHandler("start", start_command))
    app.add_handler(CommandHandler("upload", upload_command))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(CommandHandler("test", test_command))
    app.add_handler(conv_handler)

    # Регистрируем ежедневное выполнение автоотчёта (например, в 09:00)
    kiev_tz = pytz.timezone("Europe/Kiev")
    app.job_queue.run_daily(
        auto_report_job,
        time=time(hour=00, minute=6, second=0, tzinfo=kiev_tz),
        name="auto_report_job"
    )

    logging.info("Бот запущен. Ctrl+C для остановки.")
    app.run_polling()


if __name__ == "__main__":
    main()
